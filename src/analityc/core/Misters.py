import asyncio
import base64
import csv
import datetime
import logging
import os
import threading
import httpx
import uuid
import unicodedata
from collections import deque
from typing import Any, Dict, List, Optional, Tuple

import cv2
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from ultralytics import YOLO

from ..config.config import DEFAULT_ROI


logger = logging.getLogger(__name__)


DEFAULT_MODEL_PATH = r"C:\\Users\\Sistema-1\\Desktop\\ELDE\\SERVER-IA PERIMETRALES\\models\\base\\1080.pt"
DEFAULT_PERSON_MODEL_PATH = r"C:\\Users\\Sistema-1\\Desktop\\ELDE\\SERVER-IA PERIMETRALES\\models\\base\\yolo12m.pt"
DEFAULT_OCCUPY_SCREENSHOT_DIR = r"C:\\Users\\Sistema-1\\Desktop\\ELDE\\SERVER-IA PERIMETRALES\\output\\Cliente_ocupa_mesa_misters"
DEFAULT_VACATE_SCREENSHOT_DIR = r"C:\\Users\\Sistema-1\\Desktop\\ELDE\\SERVER-IA PERIMETRALES\\output\\Cliente_desocupa_mesa_misters"
DEFAULT_MANAGER_SCREENSHOT_DIR = r"C:\\Users\\Sistema-1\\Desktop\\ELDE\\SERVER-IA PERIMETRALES\\output\\Toque_de_gerente_misters"
DEFAULT_ASSISTANT_SCREENSHOT_DIR = r"C:\\Users\\Sistema-1\\Desktop\\ELDE\\SERVER-IA PERIMETRALES\\output\\Toque_de_asistente_misters"
DEFAULT_ALERT_CSV_PATH = r"C:\\Users\\Sistema-1\\Desktop\\ELDE\\SERVER-IA PERIMETRALES\\output\\misters_alertas.csv"
DEFAULT_ALERT_EXCEL_PATH = r"C:\\Users\\Sistema-1\\Desktop\\ELDE\\SERVER-IA PERIMETRALES\\output\\misters_alertas.xlsx"
DEFAULT_ALERT_PENDING_CSV_PATH = r"C:\\Users\\Sistema-1\\Desktop\\ELDE\\SERVER-IA PERIMETRALES\\output\\misters_alertas_pending.csv"


class EventConfirmationBuffer:
    """
    Ventana deslizante de confirmación temporal para un par (global_id, label).

    Un evento se confirma cuando:
      - aparece en `required_frames` de los últimos `window_size` frames
      - la confianza media de esas apariciones supera `confidence_threshold`
    """

    def __init__(self, window_size: int = 5, required_frames: int = 3, confidence_threshold: float = 0.60):
        self.window_size = window_size
        self.required_frames = required_frames
        self.confidence_threshold = confidence_threshold
        # deque of (frame_number, confidence)
        self._hits: deque = deque(maxlen=window_size)

    def add(self, frame_number: int, confidence: float) -> None:
        self._hits.append((frame_number, confidence))

    def is_confirmed(self, current_frame: int) -> bool:
        # Solo hits dentro de la ventana actual
        recent = [(f, c) for f, c in self._hits if (current_frame - f) < self.window_size]
        if len(recent) < self.required_frames:
            return False
        avg_conf = sum(c for _, c in recent) / len(recent)
        return avg_conf >= self.confidence_threshold

    def reset(self) -> None:
        self._hits.clear()


class MistersProcess:
    """
    Detector de eventos:
      - 'Cliente ocupa mesa'
      - 'Cliente desocupa mesa'
      - 'Toque de gerente'
      - 'Toque de asistente'
    usando YOLO.

    Versión v2 – Robustecida contra falsos positivos mediante:
      - Confirmación temporal multi-frame
      - Votación de confianza acumulada
      - Requisito estricto de persona asociada
      - Umbrales adaptativos por clase
      - Filtros de área y aspect-ratio
      - NMS inter-detección de eventos
      - Cooldowns configurables por tipo de evento
    """

    # Mapeo de etiquetas del modelo → nombre interno del evento
    # Ajusta las claves al nombre exacto que devuelve tu modelo YOLO
    TARGET_LABELS: Dict[str, str] = {
        "cliente ocupa mesa":    "occupy",
        "cliente desocupa mesa": "vacate",
        "toque de gerente":      "manager_touch",
        "toque de asistente":    "assistant_touch",
    }

    # Nombre legible para cada evento interno
    EVENT_DISPLAY_NAMES: Dict[str, str] = {
        "occupy":         "Cliente ocupa mesa",
        "vacate":         "Cliente desocupa mesa",
        "manager_touch":  "Toque de gerente",
        "assistant_touch": "Toque de asistente",
    }

    def __init__(
        self,
        client_id: Optional[str] = None,
        model_path: str = DEFAULT_MODEL_PATH,
        person_model_path: str = DEFAULT_PERSON_MODEL_PATH,
        confidence_threshold: float = 0.5,
        iou_threshold: float = 0.5,
        person_confidence_threshold: Optional[float] = None,
        person_iou_threshold: Optional[float] = None,
        device: str = "cpu",
        log_file: str = "output/misters_log.csv",
        image_quality: int = 85,
        server_url: Optional[str] = None,
        component_key: str = "misters_inference",
        type_inference: str = "misters",
        max_image_size: Tuple[int, int] = (640, 640),
        tracker: str = "trackers/botsort_reid.yaml",
        id_reset_frames: int = 900,
        min_person_confidence: float = 0.3,
        min_event_confidence: Optional[float] = None,
        min_person_area: int = 900,
        max_event_person_distance: float = 120.0,
        min_event_person_iou: float = 0.08,
        appearance_match_threshold: float = 0.85,
        max_reid_distance: float = 80.0,
        appearance_update_alpha: float = 0.35,
        event_cooldown_frames: int = 90,
        # Cooldowns por tipo de evento (None = usa event_cooldown_frames)
        occupy_cooldown_frames: Optional[int] = None,
        vacate_cooldown_frames: Optional[int] = None,
        manager_touch_cooldown_frames: Optional[int] = 300,
        assistant_touch_cooldown_frames: Optional[int] = 300,
        min_track_age_frames: int = 5,
        # Directorios de screenshots por tipo de evento
        occupy_screenshot_dir: str = DEFAULT_OCCUPY_SCREENSHOT_DIR,
        vacate_screenshot_dir: str = DEFAULT_VACATE_SCREENSHOT_DIR,
        manager_screenshot_dir: str = DEFAULT_MANAGER_SCREENSHOT_DIR,
        assistant_screenshot_dir: str = DEFAULT_ASSISTANT_SCREENSHOT_DIR,
        alert_csv_path: str = DEFAULT_ALERT_CSV_PATH,
        excel_autofit_interval: int = 10,
        # ── PARÁMETROS ANTI FALSOS POSITIVOS ─────────────────────────────────
        temporal_confirmation_frames: int = 1,
        temporal_window_size: int = 1,
        # Confirmación por clase (None = usa el global)
        occupy_confirmation_frames: Optional[int] = 1,
        occupy_window_size: Optional[int] = 1,
        vacate_confirmation_frames: Optional[int] = None,
        vacate_window_size: Optional[int] = None,
        manager_touch_confirmation_frames: Optional[int] = None,
        manager_touch_window_size: Optional[int] = None,
        assistant_touch_confirmation_frames: Optional[int] = None,
        assistant_touch_window_size: Optional[int] = None,
        # Umbrales de confianza por clase (None = usa min_event_confidence)
        occupy_confidence_threshold: Optional[float] = 0.01,
        vacate_confidence_threshold: Optional[float] = None,
        manager_touch_confidence_threshold: Optional[float] = None,
        assistant_touch_confidence_threshold: Optional[float] = None,
        # Votación de confianza acumulada
        confidence_vote_threshold: float = 0.01,
        # Requisito estricto de persona
        require_person_match: bool = False,
        # Filtro de área mínima del bbox de evento
        min_event_area: int = 1,
        # Filtro de aspect ratio (min_ratio, max_ratio)
        event_aspect_ratio_range: Tuple[float, float] = (0.01, 10.0),
        # NMS entre detecciones de eventos en el mismo frame
        event_nms_iou_threshold: float = 0.9,
        # Máximo de eventos distintos por frame
        max_events_per_frame: int = 10,
        # Suprimir detecciones cuya confianza cae en cada frame consecutivo
        suppress_confidence_drop: bool = False,
        confidence_drop_threshold: float = 1.0,
    ):
        self.client_id = client_id or str(uuid.uuid4())
        self.component_key = component_key
        self.type_inference = type_inference
        self.server_url = server_url

        self.confidence_threshold = confidence_threshold
        self.iou_threshold = iou_threshold
        self.model_path = self._resolve_model_path(model_path, DEFAULT_MODEL_PATH, "principal")
        self.person_model_path = self._resolve_model_path(person_model_path, DEFAULT_PERSON_MODEL_PATH, "personas")
        self.person_confidence_threshold = (
            person_confidence_threshold if person_confidence_threshold is not None else confidence_threshold
        )
        self.person_iou_threshold = person_iou_threshold if person_iou_threshold is not None else iou_threshold
        self.image_quality = image_quality
        self.max_image_size = max_image_size
        self.log_file = log_file
        self.device = device

        if tracker and not os.path.isabs(tracker):
            base_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", ".."))
            self.tracker = os.path.join(base_dir, tracker)
        else:
            self.tracker = tracker

        self.id_reset_frames = id_reset_frames
        self.min_person_confidence = min_person_confidence
        self.min_event_confidence = min_event_confidence if min_event_confidence is not None else confidence_threshold
        self.min_person_area = min_person_area
        self.max_event_person_distance = max_event_person_distance
        self.min_event_person_iou = min_event_person_iou
        self.appearance_match_threshold = appearance_match_threshold
        self.max_reid_distance = max_reid_distance
        self.appearance_update_alpha = appearance_update_alpha

        # Cooldowns: por tipo > global
        self.event_cooldown_frames = event_cooldown_frames
        self.occupy_cooldown_frames = occupy_cooldown_frames if occupy_cooldown_frames is not None else event_cooldown_frames
        self.vacate_cooldown_frames = vacate_cooldown_frames if vacate_cooldown_frames is not None else event_cooldown_frames
        self.manager_touch_cooldown_frames = (
            manager_touch_cooldown_frames if manager_touch_cooldown_frames is not None else event_cooldown_frames
        )
        self.assistant_touch_cooldown_frames = (
            assistant_touch_cooldown_frames if assistant_touch_cooldown_frames is not None else event_cooldown_frames
        )

        self.min_track_age_frames = min_track_age_frames

        # Directorios de screenshots
        self.occupy_screenshot_dir = occupy_screenshot_dir
        self.vacate_screenshot_dir = vacate_screenshot_dir
        self.manager_screenshot_dir = manager_screenshot_dir
        self.assistant_screenshot_dir = assistant_screenshot_dir

        self.alert_csv_path = alert_csv_path
        self.alert_excel_path = DEFAULT_ALERT_EXCEL_PATH
        self.alert_pending_csv_path = DEFAULT_ALERT_PENDING_CSV_PATH
        self.excel_autofit_interval = max(1, int(excel_autofit_interval))
        self._excel_alert_write_count = 0

        # ── Parámetros anti FP ────────────────────────────────────────────────
        self.temporal_confirmation_frames = max(1, temporal_confirmation_frames)
        self.temporal_window_size = max(self.temporal_confirmation_frames + 1, temporal_window_size)

        # Confirmación temporal por clase
        self.occupy_confirmation_frames = max(1, occupy_confirmation_frames if occupy_confirmation_frames is not None else self.temporal_confirmation_frames)
        self.occupy_window_size = max(self.occupy_confirmation_frames, occupy_window_size if occupy_window_size is not None else self.temporal_window_size)

        self.vacate_confirmation_frames = max(1, vacate_confirmation_frames if vacate_confirmation_frames is not None else self.temporal_confirmation_frames)
        self.vacate_window_size = max(self.vacate_confirmation_frames + 1, vacate_window_size if vacate_window_size is not None else self.temporal_window_size)

        self.manager_touch_confirmation_frames = max(1, manager_touch_confirmation_frames if manager_touch_confirmation_frames is not None else self.temporal_confirmation_frames)
        self.manager_touch_window_size = max(self.manager_touch_confirmation_frames + 1, manager_touch_window_size if manager_touch_window_size is not None else self.temporal_window_size)

        self.assistant_touch_confirmation_frames = max(1, assistant_touch_confirmation_frames if assistant_touch_confirmation_frames is not None else self.temporal_confirmation_frames)
        self.assistant_touch_window_size = max(self.assistant_touch_confirmation_frames + 1, assistant_touch_window_size if assistant_touch_window_size is not None else self.temporal_window_size)

        # Umbrales de confianza por clase
        self.occupy_confidence_threshold = (
            occupy_confidence_threshold if occupy_confidence_threshold is not None else 0.45
        )
        self.vacate_confidence_threshold = (
            vacate_confidence_threshold if vacate_confidence_threshold is not None else self.min_event_confidence
        )
        self.manager_touch_confidence_threshold = (
            manager_touch_confidence_threshold if manager_touch_confidence_threshold is not None else self.min_event_confidence
        )
        self.assistant_touch_confidence_threshold = (
            assistant_touch_confidence_threshold if assistant_touch_confidence_threshold is not None else self.min_event_confidence
        )

        self.confidence_vote_threshold = confidence_vote_threshold
        self.require_person_match = require_person_match
        self.min_event_area = min_event_area
        self.event_aspect_ratio_range = event_aspect_ratio_range
        self.event_nms_iou_threshold = event_nms_iou_threshold
        self.max_events_per_frame = max_events_per_frame
        self.suppress_confidence_drop = suppress_confidence_drop
        self.confidence_drop_threshold = confidence_drop_threshold
        self._occupy_relax_filters = True

        # Buffers de confirmación temporal: {(global_id, label): EventConfirmationBuffer}
        self._confirmation_buffers: Dict[Tuple[Any, str], EventConfirmationBuffer] = {}
        # Última confianza por (global_id, label) para detectar caídas
        self._last_confidence: Dict[Tuple[Any, str], float] = {}

        # Contadores por tipo de evento
        self.occupy_count = 0
        self.vacate_count = 0
        self.manager_touch_count = 0
        self.assistant_touch_count = 0
        self.total_events = 0
        self.frame_counter = 0
        self.frame_events: List[Dict[str, Any]] = []

        self.roi_polygon = np.array(DEFAULT_ROI, np.int32)
        self.roi_polygon_points = self.roi_polygon.reshape((-1, 1, 2))

        self.event_history: Dict[Any, set] = {}
        self.unique_track_ids: set = set()
        self.track_id_map: Dict[int, int] = {}
        self.track_last_seen: Dict[int, int] = {}
        self.next_global_id = 1
        self.last_processed_frame: Optional[np.ndarray] = None
        self.global_id_last_seen: Dict[int, int] = {}
        self.global_id_last_center: Dict[int, Tuple[float, float]] = {}
        self.global_id_features: Dict[int, np.ndarray] = {}
        self.event_last_seen: Dict[Tuple[Any, str], int] = {}
        self.global_id_first_seen: Dict[int, int] = {}
        self.saved_screenshot_keys: set = set()
        self.saved_screenshot_paths: Dict[Tuple[str, Any], str] = {}
        self.active_persons: Dict[int, Tuple[float, float]] = {}
        self.has_person_class = False

        # Deduplicación espacial: (label) -> [(frame, cx, cy), ...]
        self._spatial_event_log: Dict[str, List[Tuple[int, float, float]]] = {
            "occupy": [], "vacate": [], "manager_touch": [], "assistant_touch": []
        }
        self._spatial_dedup_distance = 150.0
        self._spatial_dedup_frames = 300

        self.target_labels = {self._normalize_label(k): v for k, v in self.TARGET_LABELS.items()}
        self.person_labels = {"person", "persona", "person/persona"}
        self.person_labels_normalized = {self._normalize_label(x) for x in self.person_labels}

        self.model = None
        self.person_model = None
        self._initialize_model()
        self.setup_log_file()

        logger.info(
            "MistersProcess v2 listo. Anti-FP activo: confirmación=%d/%d frames, "
            "require_person=%s, min_event_area=%d",
            self.temporal_confirmation_frames,
            self.temporal_window_size,
            self.require_person_match,
            self.min_event_area,
        )

    # ──────────────────────────────────────────────────────────────────────────
    # Infraestructura
    # ──────────────────────────────────────────────────────────────────────────

    def _resolve_model_path(self, model_path: str, fallback_path: str, model_kind: str) -> str:
        candidates: List[str] = []
        if model_path:
            candidates.append(model_path)
        base_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", ".."))
        if model_path and not os.path.isabs(model_path):
            candidates.append(os.path.join(base_dir, model_path))
            candidates.append(os.path.join(base_dir, "models", "base", model_path))
        seen: set = set()
        for candidate in candidates:
            normalized = os.path.abspath(candidate)
            if normalized in seen:
                continue
            seen.add(normalized)
            if os.path.isfile(normalized):
                return normalized
        fallback = os.path.abspath(fallback_path) if fallback_path else ""
        if fallback and os.path.isfile(fallback):
            logger.warning("Modelo %s no encontrado en '%s'. Usando fallback '%s'.", model_kind, model_path, fallback)
            return fallback
        tried = ", ".join(os.path.abspath(p) for p in candidates) or str(model_path)
        raise FileNotFoundError(
            f"No se encontró el modelo {model_kind}. Rutas probadas: {tried}. Fallback ausente: {fallback or 'N/A'}"
        )

    def _initialize_model(self):
        try:
            self.model = YOLO(self.model_path).to(self.device)
            self.person_model = YOLO(self.person_model_path).to(self.device)
            person_names = {self._normalize_label(name) for name in self.person_model.names.values()}
            self.has_person_class = any(name in self.person_labels_normalized for name in person_names)
            logger.info("Modelo YOLO cargado en %s", self.device)
        except Exception as exc:
            logger.error("No se pudo cargar el modelo: %s", exc)
            raise

    def setup_log_file(self):
        try:
            if self.log_file:
                os.makedirs(os.path.dirname(self.log_file), exist_ok=True)
                with open(self.log_file, "w", encoding="utf-8") as f:
                    f.write("timestamp,frame,occupy,vacate,manager_touch,assistant_touch,total,event\n")
        except Exception:
            logger.warning("No se pudo inicializar el log %s", self.log_file)
        for d in [
            self.occupy_screenshot_dir,
            self.vacate_screenshot_dir,
            self.manager_screenshot_dir,
            self.assistant_screenshot_dir,
        ]:
            try:
                os.makedirs(d, exist_ok=True)
            except Exception as exc:
                logger.warning("No se pudo crear carpeta de screenshots '%s': %s", d, exc)
        self._setup_alert_csv()
        self._setup_alert_excel()

    def _setup_alert_csv(self) -> None:
        try:
            os.makedirs(os.path.dirname(self.alert_csv_path), exist_ok=True)
            if not os.path.isfile(self.alert_csv_path):
                with open(self.alert_csv_path, "w", newline="", encoding="utf-8") as csv_file:
                    writer = csv.writer(csv_file)
                    writer.writerow([
                        "occupy", "vacate", "manager_touch", "assistant_touch",
                        "occupy_id", "vacate_id", "manager_touch_id", "assistant_touch_id",
                    ])
        except Exception as exc:
            logger.warning("No se pudo inicializar CSV de alertas %s: %s", self.alert_csv_path, exc)

    def _setup_alert_excel(self) -> None:
        try:
            os.makedirs(os.path.dirname(self.alert_excel_path), exist_ok=True)
            expected_headers = [
                "Cliente ocupa mesa", "Foto ocupa",
                "Cliente desocupa mesa", "Foto desocupa",
                "Toque de gerente", "Foto gerente",
                "Toque de asistente", "Foto asistente",
            ]
            if not os.path.isfile(self.alert_excel_path):
                workbook = Workbook()
                sheet = workbook.active
                sheet.title = "alertas"
                sheet.append(expected_headers)
                workbook.save(self.alert_excel_path)
                workbook.close()
            else:
                workbook = load_workbook(self.alert_excel_path)
                sheet = workbook.active
                current_headers = [sheet.cell(row=1, column=i).value for i in range(1, len(expected_headers) + 1)]
                if current_headers != expected_headers:
                    sheet.delete_rows(1, sheet.max_row)
                    sheet.append(expected_headers)
                workbook.save(self.alert_excel_path)
                workbook.close()
        except Exception as exc:
            logger.warning("No se pudo inicializar Excel de alertas %s: %s", self.alert_excel_path, exc)

    # Mapping: event label → (csv_col_index, excel_col_index, excel_photo_col)
    _EVENT_CSV_EXCEL_MAP: Dict[str, Tuple[int, int, str]] = {
        "occupy":         (0, 1, "B"),
        "vacate":         (2, 3, "D"),
        "manager_touch":  (4, 5, "F"),
        "assistant_touch": (6, 7, "H"),
    }

    def _build_alert_row(self, event: Dict[str, Any]) -> Optional[List[str]]:
        event_name = str(event.get("event", "")).strip().lower()
        if event_name not in self._EVENT_CSV_EXCEL_MAP:
            return None
        timestamp = self._format_timestamp_millis()
        event_id = str(event.get("global_id", event.get("track_id", "na")))
        row = ["", "", "", "", "", "", "", ""]
        col_idx, _, _ = self._EVENT_CSV_EXCEL_MAP[event_name]
        row[col_idx] = timestamp
        row[col_idx + 1] = event_id
        return row

    def _build_excel_row(self, event: Dict[str, Any]) -> Optional[Tuple[List[str], str]]:
        event_name = str(event.get("event", "")).strip().lower()
        if event_name not in self._EVENT_CSV_EXCEL_MAP:
            return None
        timestamp = self._format_timestamp_millis()
        _, excel_col, photo_col = self._EVENT_CSV_EXCEL_MAP[event_name]
        row = ["", "", "", "", "", "", "", ""]
        row[excel_col - 1] = timestamp
        return row, photo_col

    def _append_row_to_csv_file(self, file_path: str, row: List[str]) -> None:
        with open(file_path, "a", newline="", encoding="utf-8") as csv_file:
            writer = csv.writer(csv_file)
            writer.writerow(row)

    def _flush_pending_alerts(self) -> None:
        if not os.path.isfile(self.alert_pending_csv_path):
            return
        try:
            with open(self.alert_pending_csv_path, "r", newline="", encoding="utf-8") as pending_file:
                rows = list(csv.reader(pending_file))
            if not rows:
                os.remove(self.alert_pending_csv_path)
                return
            with open(self.alert_csv_path, "a", newline="", encoding="utf-8") as main_file:
                writer = csv.writer(main_file)
                writer.writerows(rows)
            os.remove(self.alert_pending_csv_path)
        except PermissionError:
            return
        except Exception as exc:
            logger.warning("No se pudo volcar CSV pendiente de alertas: %s", exc)

    def _append_alert_to_excel(self, row: List[str], screenshot_path: str = "", preview_column: str = "B") -> None:
        try:
            if not os.path.isfile(self.alert_excel_path):
                self._setup_alert_excel()
            workbook = load_workbook(self.alert_excel_path)
            sheet = workbook.active
            sheet.append(row)
            row_idx = sheet.max_row
            self._insert_image_preview(sheet, row_idx, screenshot_path, preview_column)
            self._excel_alert_write_count += 1
            if self._excel_alert_write_count % self.excel_autofit_interval == 0:
                self._autofit_excel_columns(sheet)
            workbook.save(self.alert_excel_path)
            workbook.close()
        except PermissionError:
            logger.warning("Excel de alertas bloqueado: %s", self.alert_excel_path)
        except Exception as exc:
            logger.warning("No se pudo registrar alerta en Excel: %s", exc)

    def _autofit_excel_columns(self, sheet) -> None:
        try:
            for col_idx, column_cells in enumerate(sheet.iter_cols(min_row=1, max_row=sheet.max_row), start=1):
                max_len = max((len(str(cell.value or "")) for cell in column_cells), default=0)
                sheet.column_dimensions[get_column_letter(col_idx)].width = max(12, max_len + 2)
        except Exception:
            return

    def _insert_image_preview(self, sheet, row_idx: int, screenshot_path: str, preview_column: str) -> None:
        if not screenshot_path or not os.path.isfile(screenshot_path):
            return
        try:
            img = XLImage(screenshot_path)
            img.width, img.height = 220, 150
            sheet.row_dimensions[row_idx].height = 115
            sheet.column_dimensions[preview_column].width = max(sheet.column_dimensions[preview_column].width or 0, 35)
            sheet.add_image(img, f"{preview_column}{row_idx}")
        except Exception as exc:
            logger.warning("No se pudo insertar preview de imagen en Excel: %s", exc)

    def _format_timestamp_millis(self, raw_timestamp: Any = None) -> str:
        dt_value: datetime.datetime
        if raw_timestamp is None:
            dt_value = datetime.datetime.now()
        elif isinstance(raw_timestamp, datetime.datetime):
            dt_value = raw_timestamp
        elif raw_timestamp:
            raw_str = str(raw_timestamp).strip()
            for fmt in ("%d/%m/%Y %H:%M:%S.%f", "%d/%m/%Y %H:%M:%S"):
                try:
                    dt_value = datetime.datetime.strptime(raw_str, fmt)
                    break
                except ValueError:
                    pass
            else:
                try:
                    dt_value = datetime.datetime.fromisoformat(raw_str.replace("Z", "+00:00"))
                except ValueError:
                    dt_value = datetime.datetime.now()
        else:
            dt_value = datetime.datetime.now()
        return dt_value.strftime("%d/%m/%Y %H:%M:%S.%f")[:-3]

    def _append_alert_to_csv(self, event: Dict[str, Any]) -> None:
        csv_row = self._build_alert_row(event)
        excel_data = self._build_excel_row(event)
        if csv_row is None or excel_data is None:
            return
        excel_row, preview_column = excel_data
        self._flush_pending_alerts()
        try:
            self._append_row_to_csv_file(self.alert_csv_path, csv_row)
        except PermissionError:
            try:
                self._append_row_to_csv_file(self.alert_pending_csv_path, csv_row)
                logger.warning("CSV principal bloqueado. Alerta guardada en pendiente: %s", self.alert_pending_csv_path)
            except Exception as exc:
                logger.warning("No se pudo registrar alerta en CSV pendiente: %s", exc)
        except Exception as exc:
            logger.warning("No se pudo registrar alerta en CSV principal: %s", exc)
        screenshot_path = str(event.get("screenshot_path", ""))
        self._append_alert_to_excel(excel_row, screenshot_path, preview_column)

    def _get_screenshot_dir(self, event_name: str) -> str:
        return {
            "occupy":         self.occupy_screenshot_dir,
            "vacate":         self.vacate_screenshot_dir,
            "manager_touch":  self.manager_screenshot_dir,
            "assistant_touch": self.assistant_screenshot_dir,
        }.get(event_name, self.occupy_screenshot_dir)

    def _get_screenshot_prefix(self, event_name: str) -> str:
        return {
            "occupy":         "cliente_ocupa_mesa",
            "vacate":         "cliente_desocupa_mesa",
            "manager_touch":  "toque_de_gerente",
            "assistant_touch": "toque_de_asistente",
        }.get(event_name, event_name)

    def _save_event_screenshot(self, frame: np.ndarray, event: Dict[str, Any]) -> Optional[str]:
        event_name = str(event.get("event", "")).strip().lower()
        target_dir = self._get_screenshot_dir(event_name)
        prefix = self._get_screenshot_prefix(event_name)
        os.makedirs(target_dir, exist_ok=True)
        track_value = event.get("global_id", event.get("track_id", "na"))
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S_%f")
        filename = f"{prefix}_frame{self.frame_counter}_id{track_value}_{timestamp}.jpg"
        file_path = os.path.join(target_dir, filename)
        ok = cv2.imwrite(file_path, frame, [cv2.IMWRITE_JPEG_QUALITY, int(self.image_quality)])
        if ok:
            logger.info("Screenshot guardado (%s): %s", event_name, file_path)
            return file_path
        logger.warning("No se pudo guardar screenshot (%s): %s", event_name, file_path)
        return None

    # ──────────────────────────────────────────────────────────────────────────
    # Geometría
    # ──────────────────────────────────────────────────────────────────────────

    def is_inside_roi(self, point: Tuple[float, float]) -> bool:
        x, y = float(point[0]), float(point[1])
        return cv2.pointPolygonTest(self.roi_polygon_points, (int(x), int(y)), False) >= 0

    def compress_image(self, image: np.ndarray) -> np.ndarray:
        try:
            h, w = image.shape[:2]
            max_w, max_h = self.max_image_size
            if w > max_w or h > max_h:
                aspect = w / h
                if w > max_w:
                    w, h = max_w, int(max_w / aspect)
                if h > max_h:
                    h, w = max_h, int(max_h * aspect)
                image = cv2.resize(image, (w, h), interpolation=cv2.INTER_AREA)
            return image
        except Exception:
            return image

    def convert_to_serializable(self, obj):
        if isinstance(obj, np.integer):
            return int(obj)
        if isinstance(obj, np.floating):
            return float(obj)
        if isinstance(obj, np.ndarray):
            return obj.tolist()
        if isinstance(obj, dict):
            return {k: self.convert_to_serializable(v) for k, v in obj.items()}
        if isinstance(obj, (list, tuple)):
            converted = [self.convert_to_serializable(x) for x in obj]
            return type(obj)(converted)
        return obj

    def _normalize_label(self, label: str) -> str:
        normalized = unicodedata.normalize("NFKD", label)
        normalized = "".join(ch for ch in normalized if not unicodedata.combining(ch))
        normalized = normalized.replace("_", " ").strip().lower()
        return " ".join(normalized.split())

    def _box_area(self, box: np.ndarray) -> float:
        x1, y1, x2, y2 = box
        return max(0.0, float(x2 - x1)) * max(0.0, float(y2 - y1))

    def _box_iou(self, box_a: np.ndarray, box_b: np.ndarray) -> float:
        ax1, ay1, ax2, ay2 = box_a
        bx1, by1, bx2, by2 = box_b
        ix1, iy1 = max(ax1, bx1), max(ay1, by1)
        ix2, iy2 = min(ax2, bx2), min(ay2, by2)
        inter = max(0.0, ix2 - ix1) * max(0.0, iy2 - iy1)
        if inter <= 0:
            return 0.0
        union = self._box_area(box_a) + self._box_area(box_b) - inter
        return float(inter / union) if union > 0 else 0.0

    def _center_distance(self, c1: Tuple[float, float], c2: Tuple[float, float]) -> float:
        return float(((c1[0] - c2[0]) ** 2 + (c1[1] - c2[1]) ** 2) ** 0.5)

    # ──────────────────────────────────────────────────────────────────────────
    # FILTROS ANTI FALSOS POSITIVOS
    # ──────────────────────────────────────────────────────────────────────────

    def _passes_geometry_filters(self, box: np.ndarray, label: str) -> bool:
        area = self._box_area(box)
        if label == "occupy" and self._occupy_relax_filters:
            return True
        if area < self.min_event_area:
            logger.debug("FP-GUARD area: label=%s area=%.0f < min=%d → descartado", label, area, self.min_event_area)
            return False
        x1, y1, x2, y2 = box
        w, h = float(x2 - x1), float(y2 - y1)
        if h <= 0:
            return False
        ratio = w / h
        min_r, max_r = self.event_aspect_ratio_range
        if not (min_r <= ratio <= max_r):
            logger.debug(
                "FP-GUARD aspect: label=%s ratio=%.2f not in [%.2f, %.2f] → descartado",
                label, ratio, min_r, max_r,
            )
            return False
        return True

    def _is_spatially_duplicate(self, label: str, center: Tuple[float, float]) -> bool:
        log = self._spatial_event_log.get(label, [])
        cutoff = self.frame_counter - self._spatial_dedup_frames
        log[:] = [(f, cx, cy) for f, cx, cy in log if f >= cutoff]
        cx, cy = center
        for _, ex, ey in log:
            dist = ((cx - ex) ** 2 + (cy - ey) ** 2) ** 0.5
            if dist < self._spatial_dedup_distance:
                return True
        return False

    def _passes_confidence_threshold(self, label: str, confidence: float) -> bool:
        if label == "occupy" and self._occupy_relax_filters:
            return True
        thresholds = {
            "occupy":         self.occupy_confidence_threshold,
            "vacate":         self.vacate_confidence_threshold,
            "manager_touch":  self.manager_touch_confidence_threshold,
            "assistant_touch": self.assistant_touch_confidence_threshold,
        }
        threshold = thresholds.get(label, self.min_event_confidence)
        if confidence < threshold:
            logger.debug("FP-GUARD conf: label=%s conf=%.3f < threshold=%.3f → descartado", label, confidence, threshold)
            return False
        return True

    def _apply_event_nms(self, candidates: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        if len(candidates) <= 1:
            return candidates
        candidates_sorted = sorted(candidates, key=lambda d: d["confidence"], reverse=True)
        kept = []
        suppressed = set()
        for i, det in enumerate(candidates_sorted):
            if i in suppressed:
                continue
            kept.append(det)
            for j in range(i + 1, len(candidates_sorted)):
                if j in suppressed:
                    continue
                iou = self._box_iou(det["box"], candidates_sorted[j]["box"])
                if iou > self.event_nms_iou_threshold:
                    suppressed.add(j)
                    logger.debug("FP-GUARD NMS: suprimido candidato j=%d (iou=%.2f con i=%d)", j, iou, i)
        return kept

    def _get_confirmation_buffer(self, global_id: Any, label: str) -> EventConfirmationBuffer:
        key = (global_id, label)
        if key not in self._confirmation_buffers:
            req_frames_map = {
                "occupy":         self.occupy_confirmation_frames,
                "vacate":         self.vacate_confirmation_frames,
                "manager_touch":  self.manager_touch_confirmation_frames,
                "assistant_touch": self.assistant_touch_confirmation_frames,
            }
            win_size_map = {
                "occupy":         self.occupy_window_size,
                "vacate":         self.vacate_window_size,
                "manager_touch":  self.manager_touch_window_size,
                "assistant_touch": self.assistant_touch_window_size,
            }
            req_frames = req_frames_map.get(label, self.temporal_confirmation_frames)
            win_size = win_size_map.get(label, self.temporal_window_size)
            self._confirmation_buffers[key] = EventConfirmationBuffer(
                window_size=win_size,
                required_frames=req_frames,
                confidence_threshold=self.confidence_vote_threshold,
            )
        return self._confirmation_buffers[key]

    def _update_confirmation(self, global_id: Any, label: str, confidence: float) -> bool:
        if global_id is not None and global_id in self.event_history:
            if label in self.event_history[global_id]:
                return False

        buf = self._get_confirmation_buffer(global_id, label)

        if self.suppress_confidence_drop:
            key = (global_id, label)
            prev_conf = self._last_confidence.get(key)
            if prev_conf is not None and (prev_conf - confidence) > self.confidence_drop_threshold:
                logger.debug(
                    "FP-GUARD conf-drop: id=%s label=%s prev=%.3f cur=%.3f → reset buffer",
                    global_id, label, prev_conf, confidence,
                )
                buf.reset()
            self._last_confidence[key] = confidence

        buf.add(self.frame_counter, confidence)
        confirmed = buf.is_confirmed(self.frame_counter)
        if not confirmed:
            logger.debug(
                "FP-GUARD temporal: id=%s label=%s frame=%d → pendiente confirmación",
                global_id, label, self.frame_counter,
            )
        return confirmed

    def _get_cooldown_frames(self, label: str) -> int:
        return {
            "occupy":         self.occupy_cooldown_frames,
            "vacate":         self.vacate_cooldown_frames,
            "manager_touch":  self.manager_touch_cooldown_frames,
            "assistant_touch": self.assistant_touch_cooldown_frames,
        }.get(label, self.event_cooldown_frames)

    # ──────────────────────────────────────────────────────────────────────────
    # ReID y tracking
    # ──────────────────────────────────────────────────────────────────────────

    def _extract_appearance_signature(self, image: np.ndarray, box: np.ndarray) -> Optional[np.ndarray]:
        try:
            x1, y1, x2, y2 = [int(max(0, v)) for v in box]
            h, w = image.shape[:2]
            x1, x2 = max(0, x1), min(w, x2)
            y1, y2 = max(0, y1), min(h, y2)
            if x2 <= x1 or y2 <= y1:
                return None
            crop = image[y1:y2, x1:x2]
            if crop.size == 0:
                return None
            ch, cw = crop.shape[:2]
            upper = crop[: int(ch * 0.55), :]
            lower = crop[int(ch * 0.55):, :]
            hsv_upper = cv2.cvtColor(upper, cv2.COLOR_BGR2HSV)
            hsv_lower = cv2.cvtColor(lower, cv2.COLOR_BGR2HSV)
            hist_upper = cv2.normalize(
                cv2.calcHist([hsv_upper], [0, 1], None, [24, 24], [0, 180, 0, 256]), None
            ).flatten()
            hist_lower = cv2.normalize(
                cv2.calcHist([hsv_lower], [0, 1], None, [24, 24], [0, 180, 0, 256]), None
            ).flatten()
            lab_upper = cv2.cvtColor(upper, cv2.COLOR_BGR2LAB)
            lab_lower = cv2.cvtColor(lower, cv2.COLOR_BGR2LAB)
            upper_mean = np.array(cv2.mean(lab_upper)[:3], dtype=np.float32) / 255.0
            lower_mean = np.array(cv2.mean(lab_lower)[:3], dtype=np.float32) / 255.0
            ycrcb = cv2.cvtColor(crop, cv2.COLOR_BGR2YCrCb)
            skin_mask = cv2.inRange(ycrcb, (0, 133, 77), (255, 173, 127))
            skin_ratio = float(np.count_nonzero(skin_mask)) / float(skin_mask.size) if skin_mask.size else 0.0
            skin_mean_bgr = cv2.mean(crop, mask=skin_mask)[:3]
            skin_patch = np.full((4, 4, 3), skin_mean_bgr, dtype=np.uint8)
            skin_lab = cv2.cvtColor(skin_patch, cv2.COLOR_BGR2LAB)
            skin_mean = (np.array(cv2.mean(skin_lab)[:3], dtype=np.float32) / 255.0).tolist()
            signature = np.concatenate(
                [hist_upper, hist_lower, upper_mean, lower_mean,
                 np.array([skin_ratio, skin_mean[0], skin_mean[1], skin_mean[2]])]
            ).astype(np.float32)
            norm = np.linalg.norm(signature)
            if norm > 0:
                signature /= norm
            return signature
        except Exception:
            return None

    def _compare_features(self, f1: np.ndarray, f2: np.ndarray) -> float:
        denom = float(np.linalg.norm(f1) * np.linalg.norm(f2))
        if denom <= 0:
            return 0.0
        return float(np.dot(f1, f2) / denom)

    def _update_global_features(self, global_id: int, feature: Optional[np.ndarray], center: Tuple[float, float]) -> None:
        self.global_id_last_seen[global_id] = self.frame_counter
        self.global_id_last_center[global_id] = center
        self.global_id_first_seen.setdefault(global_id, self.frame_counter)
        if feature is None:
            return
        existing = self.global_id_features.get(global_id)
        if existing is None:
            self.global_id_features[global_id] = feature
        else:
            alpha = self.appearance_update_alpha
            blended = (1 - alpha) * existing + alpha * feature
            norm = np.linalg.norm(blended)
            if norm > 0:
                blended /= norm
            self.global_id_features[global_id] = blended

    def _assign_global_id(
        self,
        track_id: Optional[int],
        center: Tuple[float, float],
        feature: Optional[np.ndarray],
        used_ids: set,
    ) -> Optional[int]:
        if track_id is not None:
            existing = self.track_id_map.get(track_id)
            if existing is not None:
                self.track_last_seen[track_id] = self.frame_counter
                self._update_global_features(existing, feature, center)
                used_ids.add(existing)
                return existing
        best_id, best_score = None, 0.0
        for gid, feat in self.global_id_features.items():
            if gid in used_ids:
                continue
            last_seen = self.global_id_last_seen.get(gid, 0)
            if self.id_reset_frames > 0 and (self.frame_counter - last_seen) > self.id_reset_frames:
                continue
            last_center = self.global_id_last_center.get(gid)
            if last_center is None or self._center_distance(center, last_center) > self.max_reid_distance:
                continue
            if feature is None or feat is None:
                continue
            score = self._compare_features(feature, feat)
            if score > best_score:
                best_score, best_id = score, gid
        if best_id is not None and best_score >= self.appearance_match_threshold:
            if track_id is not None:
                self.track_id_map[track_id] = best_id
                self.track_last_seen[track_id] = self.frame_counter
            self._update_global_features(best_id, feature, center)
            used_ids.add(best_id)
            return best_id
        if track_id is None:
            return None
        global_id = self.next_global_id
        self.next_global_id += 1
        self.track_id_map[track_id] = global_id
        self.track_last_seen[track_id] = self.frame_counter
        self._update_global_features(global_id, feature, center)
        used_ids.add(global_id)
        return global_id

    def _release_track(self, track_id: int) -> None:
        global_id = self.track_id_map.pop(track_id, None)
        self.track_last_seen.pop(track_id, None)
        if global_id is not None:
            self.global_id_last_seen[global_id] = self.frame_counter

    def _expire_stale_tracks(self) -> None:
        if self.id_reset_frames <= 0:
            return
        stale_tracks = [
            tid for tid, last in self.track_last_seen.items()
            if (self.frame_counter - last) > self.id_reset_frames
        ]
        for tid in stale_tracks:
            self._release_track(tid)
        stale_global = [
            gid for gid, last in self.global_id_last_seen.items()
            if (self.frame_counter - last) > self.id_reset_frames
        ]
        for gid in stale_global:
            self.global_id_last_seen.pop(gid, None)
            self.global_id_last_center.pop(gid, None)
            self.global_id_features.pop(gid, None)
            self.global_id_first_seen.pop(gid, None)
            self.event_history.pop(gid, None)
            for lbl in self.TARGET_LABELS.values():
                self.event_last_seen.pop((gid, lbl), None)
                self._confirmation_buffers.pop((gid, lbl), None)
                self._last_confidence.pop((gid, lbl), None)

    def _is_track_stable(self, global_id: Optional[int]) -> bool:
        if global_id is None:
            return True
        if self._occupy_relax_filters:
            return True
        first_seen = self.global_id_first_seen.get(global_id)
        if first_seen is None:
            return False
        return (self.frame_counter - first_seen) >= self.min_track_age_frames

    # ──────────────────────────────────────────────────────────────────────────
    # Registro de eventos
    # ──────────────────────────────────────────────────────────────────────────

    def _register_event(self, label: str, track_id: Optional[int], center: Tuple[float, float]) -> Optional[Dict[str, Any]]:
        cooldown = self._get_cooldown_frames(label)
        if cooldown > 0 and track_id is not None:
            last_seen = self.event_last_seen.get((track_id, label))
            if last_seen is not None and (self.frame_counter - last_seen) <= cooldown:
                return None
        if track_id is not None:
            if track_id in self.event_history and label in self.event_history[track_id]:
                return None
            self.event_history.setdefault(track_id, set()).add(label)
        else:
            cx, cy = int(center[0]), int(center[1])
            key = f"no_track:{cx}:{cy}"
            if key in self.event_history and label in self.event_history[key]:
                return None
            self.event_history.setdefault(key, set()).add(label)
        if track_id is not None:
            self.event_last_seen[(track_id, label)] = self.frame_counter

        # Incrementar contador del tipo de evento correspondiente
        if label == "occupy":
            self.occupy_count += 1
        elif label == "vacate":
            self.vacate_count += 1
        elif label == "manager_touch":
            self.manager_touch_count += 1
        elif label == "assistant_touch":
            self.assistant_touch_count += 1
        self.total_events += 1

        return {
            "event": label,
            "track_id": track_id,
            "center": [float(center[0]), float(center[1])],
            "timestamp": datetime.datetime.now().isoformat(),
            "frame": int(self.frame_counter),
        }

    def _store_frame_counts(self, frame_number, occupy, vacate, manager_touch, assistant_touch, total, events, tracks_in_roi) -> None:
        self.frame_events.append({
            "frame": frame_number,
            "timestamp": datetime.datetime.now().isoformat(),
            "occupy": occupy,
            "vacate": vacate,
            "manager_touch": manager_touch,
            "assistant_touch": assistant_touch,
            "total": total,
            "tracks_in_roi": tracks_in_roi,
            "events": events,
        })

    # ──────────────────────────────────────────────────────────────────────────
    # Visualización
    # ──────────────────────────────────────────────────────────────────────────

    def _draw(self, image: np.ndarray, detections: List[Dict[str, Any]]) -> np.ndarray:
        colors = {
            "occupy":         (0, 200, 255),
            "vacate":         (0, 100, 255),
            "manager_touch":  (0, 255, 128),
            "assistant_touch": (255, 200, 0),
            "person":         (200, 200, 200),
        }
        out = image.copy()
        for det in detections:
            x1, y1, x2, y2 = [int(v) for v in det["box"]]
            label = det["label"]
            color = colors.get(label, (255, 255, 255))
            cv2.rectangle(out, (x1, y1), (x2, y2), color, 2)
            display_text = self.EVENT_DISPLAY_NAMES.get(label, "Persona")
            display_id = det.get("global_id") if det.get("global_id") is not None else det.get("track_id")
            if display_id is not None:
                display_text += f" ID:{display_id}"
            cv2.putText(out, display_text, (x1, max(12, y1 - 6)), cv2.FONT_HERSHEY_SIMPLEX, 0.5, color, 2)
        overlay = out.copy()
        cv2.fillPoly(overlay, [self.roi_polygon], (0, 255, 255))
        cv2.addWeighted(overlay, 0.15, out, 0.85, 0, out)
        cv2.polylines(out, [self.roi_polygon], isClosed=True, color=(0, 255, 255), thickness=2)
        panel = out.copy()
        cv2.rectangle(panel, (5, 5), (380, 145), (0, 0, 0), -1)
        cv2.addWeighted(panel, 0.55, out, 0.45, 0, out)
        for text, pos in [
            (f"Frame: {self.frame_counter}",                     (15, 30)),
            (f"Cliente ocupa mesa: {self.occupy_count}",         (15, 55)),
            (f"Cliente desocupa mesa: {self.vacate_count}",      (15, 80)),
            (f"Toque de gerente: {self.manager_touch_count}",    (15, 105)),
            (f"Toque de asistente: {self.assistant_touch_count}",(15, 130)),
        ]:
            cv2.putText(out, text, pos, cv2.FONT_HERSHEY_SIMPLEX, 0.55, (0, 255, 255), 2)
        return out

    # ──────────────────────────────────────────────────────────────────────────
    # Red / servidor
    # ──────────────────────────────────────────────────────────────────────────

    def create_server_payload(self, frame: np.ndarray, metadata: Dict[str, Any], events: List[Dict[str, Any]]) -> Optional[Dict[str, Any]]:
        try:
            compressed = self.compress_image(frame)
            ok, buf = cv2.imencode(".jpg", compressed, [cv2.IMWRITE_JPEG_QUALITY, self.image_quality])
            if not ok:
                return None
            return {
                "id_connection": self.client_id,
                "component_key": self.component_key,
                "type_inference": self.type_inference,
                "data": {
                    "header": {
                        "frame_number": int(self.frame_counter),
                        "timestamp": datetime.datetime.now().isoformat(),
                        "event_type": "misters_events",
                        "roi_points": self.convert_to_serializable(self.roi_polygon),
                    },
                    "image": base64.b64encode(buf).decode("utf-8"),
                    "data_result": {
                        "statistics": self.convert_to_serializable(metadata),
                        "events": self.convert_to_serializable(events),
                    },
                },
            }
        except Exception as exc:
            logger.error("Error creando payload: %s", exc)
            return None

    async def send_to_server(self, payload: Dict[str, Any]) -> bool:
        if not self.server_url:
            return False
        try:
            async with httpx.AsyncClient(timeout=20.0) as client:
                resp = await client.post(self.server_url, json=payload)
                return resp.status_code == 200
        except Exception as exc:
            logger.error("Error enviando al servidor: %s", exc)
            return False

    def send_to_server_wrapper(self, payload: Dict[str, Any]):
        if not payload:
            return

        def _run():
            try:
                asyncio.run(self.send_to_server(payload))
            except RuntimeError:
                loop = asyncio.new_event_loop()
                asyncio.set_event_loop(loop)
                loop.run_until_complete(self.send_to_server(payload))
                loop.close()

        threading.Thread(target=_run, daemon=True).start()

    # ──────────────────────────────────────────────────────────────────────────
    # PROCESO PRINCIPAL DEL FRAME
    # ──────────────────────────────────────────────────────────────────────────

    def process_frame(self, image: np.ndarray, roi=None, send_to_server: bool = False) -> Tuple[np.ndarray, Dict[str, Any]]:
        self.frame_counter += 1
        if roi is not None and isinstance(roi, (list, np.ndarray)) and len(roi) >= 3:
            self.roi_polygon = np.array(roi, np.int32)
            self.roi_polygon_points = self.roi_polygon.reshape((-1, 1, 2))

        self.last_processed_frame = image.copy()
        detections: List[Dict[str, Any]] = []
        events: List[Dict[str, Any]] = []
        track_ids_in_roi: set = set()
        active_persons_this_frame: Dict[int, Tuple[float, float]] = {}
        persons_in_roi: List[Dict[str, Any]] = []
        used_global_ids: set = set()

        # Frame-level counters
        occupy_in_frame = 0
        vacate_in_frame = 0
        manager_touch_in_frame = 0
        assistant_touch_in_frame = 0

        try:
            # ── 1. Detección de personas ───────────────────────────────────────
            person_results = self.person_model.track(
                image,
                persist=True,
                conf=self.person_confidence_threshold,
                iou=self.person_iou_threshold,
                imgsz=640,
                tracker=self.tracker,
                verbose=False,
            )

            if person_results and person_results[0].boxes is not None:
                p_boxes = person_results[0].boxes
                p_xyxy = p_boxes.xyxy.cpu().numpy()
                p_cls = p_boxes.cls.cpu().numpy()
                p_conf = p_boxes.conf.cpu().numpy()
                p_ids = p_boxes.id.cpu().numpy() if p_boxes.id is not None else [None] * len(p_xyxy)

                for i in range(len(p_xyxy)):
                    cname = self._normalize_label(str(self.person_model.names[int(p_cls[i])]))
                    if cname not in self.person_labels_normalized:
                        continue
                    if float(p_conf[i]) < self.min_person_confidence:
                        continue
                    if self._box_area(p_xyxy[i]) < self.min_person_area:
                        continue
                    center = ((p_xyxy[i][0] + p_xyxy[i][2]) * 0.5, (p_xyxy[i][1] + p_xyxy[i][3]) * 0.5)
                    if not self.is_inside_roi(center):
                        raw_tid = int(p_ids[i]) if p_ids[i] is not None else None
                        if raw_tid is not None and raw_tid in self.track_id_map:
                            self._release_track(raw_tid)
                        continue

                    feature = self._extract_appearance_signature(image, p_xyxy[i])
                    raw_track_id = int(p_ids[i]) if p_ids[i] is not None else None
                    global_id = self._assign_global_id(raw_track_id, center, feature, used_global_ids)

                    det = {
                        "label": "person",
                        "box": p_xyxy[i],
                        "center": center,
                        "confidence": float(p_conf[i]),
                        "track_id": raw_track_id,
                        "global_id": global_id,
                    }
                    detections.append(det)
                    if global_id is not None:
                        track_ids_in_roi.add(global_id)
                        self.unique_track_ids.add(global_id)
                        active_persons_this_frame[global_id] = center
                        persons_in_roi.append({
                            "global_id": global_id,
                            "box": p_xyxy[i],
                            "center": center,
                            "feature": feature,
                        })

            # ── 2. Detección de eventos ────────────────────────────────────────
            event_results = self.model.predict(
                image,
                conf=self.confidence_threshold,
                iou=self.iou_threshold,
                imgsz=640,
                verbose=False,
            )

            raw_event_candidates: List[Dict[str, Any]] = []

            if event_results and event_results[0].boxes is not None:
                e_boxes = event_results[0].boxes
                e_xyxy = e_boxes.xyxy.cpu().numpy()
                e_cls = e_boxes.cls.cpu().numpy()
                e_conf = e_boxes.conf.cpu().numpy()

                for i in range(len(e_xyxy)):
                    cname = self._normalize_label(str(self.model.names[int(e_cls[i])]))
                    if cname not in self.target_labels:
                        continue
                    label = self.target_labels[cname]
                    confidence = float(e_conf[i])

                    if not self._passes_confidence_threshold(label, confidence):
                        continue

                    x1, y1, x2, y2 = e_xyxy[i]
                    center = ((x1 + x2) * 0.5, (y1 + y2) * 0.5)
                    if not self.is_inside_roi(center):
                        continue

                    if not self._passes_geometry_filters(e_xyxy[i], label):
                        continue

                    raw_event_candidates.append({
                        "label": label,
                        "box": e_xyxy[i],
                        "center": center,
                        "confidence": confidence,
                        "track_id": None,
                    })

            # NMS inter-evento
            event_candidates = self._apply_event_nms(raw_event_candidates)

            if len(event_candidates) > self.max_events_per_frame:
                event_candidates = event_candidates[: self.max_events_per_frame]
                logger.debug("FP-GUARD max_events_per_frame: recortado a %d", self.max_events_per_frame)

            for det in event_candidates:
                label = det["label"]
                confidence = det["confidence"]
                center = det["center"]

                if self.require_person_match and not persons_in_roi:
                    logger.debug("FP-GUARD no-person: label=%s frame=%d → descartado", label, self.frame_counter)
                    continue

                assigned_id = None
                if persons_in_roi:
                    best_id = None
                    best_iou = 0.0
                    best_dist = float("inf")
                    for person in persons_in_roi:
                        iou = self._box_iou(det["box"], person["box"])
                        dist = self._center_distance(center, person["center"])
                        if iou > best_iou:
                            best_iou, best_id = iou, person["global_id"]
                        if dist < best_dist:
                            best_dist = dist

                    match_by_iou = best_id is not None and best_iou >= self.min_event_person_iou
                    match_by_dist = best_dist <= self.max_event_person_distance
                    if not (match_by_iou or match_by_dist):
                        logger.debug(
                            "FP-GUARD person-dist: label=%s iou=%.2f dist=%.1f → descartado",
                            label, best_iou, best_dist,
                        )
                        continue
                    assigned_id = best_id if match_by_iou else min(
                        persons_in_roi, key=lambda p: self._center_distance(center, p["center"])
                    )["global_id"]

                if assigned_id is not None and assigned_id in self.event_history:
                    if label in self.event_history[assigned_id]:
                        continue

                if self._is_spatially_duplicate(label, center):
                    logger.debug(
                        "FP-GUARD spatial-dedup: label=%s center=(%.0f,%.0f) → duplicado espacial",
                        label, center[0], center[1],
                    )
                    continue

                if not self._is_track_stable(assigned_id):
                    logger.debug("FP-GUARD track-age: id=%s label=%s → inmaduro", assigned_id, label)
                    continue

                if not self._update_confirmation(assigned_id, label, confidence):
                    continue

                # ── Evento confirmado ─────────────────────────────────────────
                det["global_id"] = assigned_id
                detections.append(det)

                screenshot_key = (label, assigned_id)
                screenshot_path = self.saved_screenshot_paths.get(screenshot_key, "")
                if screenshot_key not in self.saved_screenshot_keys:
                    try:
                        saved = self._save_event_screenshot(
                            image, {"event": label, "global_id": assigned_id, "track_id": None}
                        )
                        if saved:
                            screenshot_path = saved
                            self.saved_screenshot_paths[screenshot_key] = saved
                        self.saved_screenshot_keys.add(screenshot_key)
                    except Exception as exc:
                        logger.warning("No se pudo guardar screenshot del evento %s: %s", label, exc)

                evt = self._register_event(label, assigned_id, center)
                if evt:
                    evt["global_id"] = assigned_id
                    evt["source_track_id"] = None
                    evt["screenshot_path"] = screenshot_path
                    events.append(evt)
                    buf_key = (assigned_id, label)
                    if buf_key in self._confirmation_buffers:
                        self._confirmation_buffers[buf_key].reset()
                    self._spatial_event_log.setdefault(label, []).append(
                        (self.frame_counter, center[0], center[1])
                    )
                    for pid in active_persons_this_frame:
                        self.event_history.setdefault(pid, set()).add(label)

            self.active_persons = active_persons_this_frame
            self._expire_stale_tracks()

        except Exception as exc:
            logger.error("Error en inferencia: %s", exc)

        processed = self._draw(image, detections)

        for evt in events:
            el = evt["event"]
            if el == "occupy":
                occupy_in_frame += 1
            elif el == "vacate":
                vacate_in_frame += 1
            elif el == "manager_touch":
                manager_touch_in_frame += 1
            elif el == "assistant_touch":
                assistant_touch_in_frame += 1

        total_frame_events = occupy_in_frame + vacate_in_frame + manager_touch_in_frame + assistant_touch_in_frame
        self._store_frame_counts(
            self.frame_counter,
            occupy_in_frame, vacate_in_frame, manager_touch_in_frame, assistant_touch_in_frame,
            total_frame_events, events, len(track_ids_in_roi),
        )

        # Convertir events a formato de alertas compatible con el cliente
        alerts_for_client = []
        for evt in events:
            event_label = evt.get("event", "")
            display_name = self.EVENT_DISPLAY_NAMES.get(event_label, event_label)

            crop_b64 = ""
            screenshot_path = evt.get("screenshot_path", "")
            if screenshot_path and os.path.isfile(screenshot_path):
                try:
                    with open(screenshot_path, "rb") as img_file:
                        crop_b64 = base64.b64encode(img_file.read()).decode("utf-8")
                except Exception as exc:
                    logger.warning("No se pudo leer screenshot para alerta: %s", exc)

            if not crop_b64 and image is not None:
                try:
                    center = evt.get("center", [])
                    if len(center) == 2:
                        cx, cy = int(center[0]), int(center[1])
                        h_img, w_img = image.shape[:2]
                        crop_size = 80
                        x1 = max(0, cx - crop_size)
                        y1 = max(0, cy - crop_size)
                        x2 = min(w_img, cx + crop_size)
                        y2 = min(h_img, cy + crop_size)
                        crop = image[y1:y2, x1:x2]
                        if crop.size > 0:
                            _, buf = cv2.imencode(".jpg", crop, [cv2.IMWRITE_JPEG_QUALITY, 75])
                            if buf is not None:
                                crop_b64 = base64.b64encode(buf).decode("utf-8")
                except Exception as exc:
                    logger.warning("No se pudo generar crop para alerta: %s", exc)

            timestamp_value = datetime.datetime.now().timestamp()
            if screenshot_path and os.path.isfile(screenshot_path):
                try:
                    timestamp_value = float(os.path.getmtime(screenshot_path))
                except Exception:
                    pass

            alerts_for_client.append({
                "event_type": display_name,
                "class_name": display_name,
                "description": f"{display_name} detectada (ID: {evt.get('global_id', evt.get('track_id', '?'))})",
                "timestamp": timestamp_value,
                "crop_image": crop_b64,
                "screenshot_path": screenshot_path or "",
            })

        metadata = {
            "frame": int(self.frame_counter),
            "occupy_count": int(self.occupy_count),
            "vacate_count": int(self.vacate_count),
            "manager_touch_count": int(self.manager_touch_count),
            "assistant_touch_count": int(self.assistant_touch_count),
            "total_events": int(self.total_events),
            "occupy_in_frame": int(occupy_in_frame),
            "vacate_in_frame": int(vacate_in_frame),
            "manager_touch_in_frame": int(manager_touch_in_frame),
            "assistant_touch_in_frame": int(assistant_touch_in_frame),
            "frame_events_total": int(total_frame_events),
            "tracks_in_roi": int(len(track_ids_in_roi)),
            "unique_tracks_total": int(len(self.unique_track_ids)),
            "roi_points": self.convert_to_serializable(self.roi_polygon),
            "alerts": alerts_for_client,
        }

        if send_to_server and self.server_url:
            payload = self.create_server_payload(processed, metadata, events)
            if payload:
                self.send_to_server_wrapper(payload)

        for evt in events:
            self._append_alert_to_csv(evt)

        if events and self.log_file:
            try:
                with open(self.log_file, "a", encoding="utf-8") as f:
                    for evt in events:
                        ts = self._format_timestamp_millis(evt.get("timestamp"))
                        f.write(
                            f"{ts},{evt['frame']},{self.occupy_count},{self.vacate_count},"
                            f"{self.manager_touch_count},{self.assistant_touch_count},"
                            f"{self.total_events},{evt['event']}\n"
                        )
            except Exception:
                pass

        return processed, {"events": events, **metadata}

    # ──────────────────────────────────────────────────────────────────────────
    # Reset y estadísticas
    # ──────────────────────────────────────────────────────────────────────────

    def reset_counter(self):
        self.occupy_count = 0
        self.vacate_count = 0
        self.manager_touch_count = 0
        self.assistant_touch_count = 0
        self.total_events = 0
        self.event_history.clear()
        self.unique_track_ids.clear()
        self.track_id_map.clear()
        self.track_last_seen.clear()
        self.next_global_id = 1
        self.frame_counter = 0
        self.global_id_last_seen.clear()
        self.global_id_last_center.clear()
        self.global_id_features.clear()
        self.event_last_seen.clear()
        self.saved_screenshot_keys.clear()
        self.saved_screenshot_paths.clear()
        self._confirmation_buffers.clear()
        self._last_confidence.clear()
        self._spatial_event_log = {
            "occupy": [], "vacate": [], "manager_touch": [], "assistant_touch": []
        }
        logger.info("Contadores reiniciados")

    def get_stats(self) -> Dict[str, Any]:
        return {
            "occupy_count": int(self.occupy_count),
            "vacate_count": int(self.vacate_count),
            "manager_touch_count": int(self.manager_touch_count),
            "assistant_touch_count": int(self.assistant_touch_count),
            "total_events": int(self.total_events),
            "frame_counter": int(self.frame_counter),
            "unique_tracks_total": int(len(self.unique_track_ids)),
            "roi_points": self.convert_to_serializable(self.roi_polygon),
        }


def create_misters_processor(**kwargs) -> MistersProcess:
    """Factory para crear instancias de MistersProcess."""
    return MistersProcess(**kwargs)